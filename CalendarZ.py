import sys  #用来执行系统程序的操作
from PyQt5 import QtWidgets as widgets
from PyQt5 import QtCore as core
from PyQt5 import QtMultimedia as media  #PyQt5用来做日历的主界面
import time  #用来查询,转换系统时间
import ntplib  #用来同步网络时间
from lunar_python import Solar, Lunar  #一个提供计算公农历历法的python库
import requests  #用来访问中国天气网，获取天气数据（默认调取北京市）
from openpyxl import load_workbook as load 
from openpyxl import Workbook  #用来读写存储备忘录的note.xlxs文件
timedifferent=0  #系统时间与网络时间的差值
birthday=None  #让用户可以输入生日
notebook=[]  #先指定notebook是列表,否则有一些BUG...
cities={'北京市':101010100,'天津市':101030100,'上海市':101020100,'重庆市':101040100,'成都市':101270101,'杭州市':101210101,'南京市':101190101,
        '深圳市':101280601,'西安市':101110101,'广州市':101280101,'青岛市':101120201,'武汉市':101200101}
choosecity='北京市'
def systemtimechange():  #与系统时间同步,差值设为0
    global timedifferent
    timedifferent=0
def nettimechange():  #与pool.ntp.org'同步,获取其与当前系统时差值
    global timedifferent
    try:
        client=ntplib.NTPClient()
        response=client.request('ntp1.aliyun.com')
        timedifferent=time.time()-response.tx_time
    except:
        widgets.QMessageBox.information(window, '同步网络时间时出错', 'CalenderZ出错了!\n这肯定不是程序的问题!\n绝对不是!\n(或许应该检查一下网络是否连接...)')
        timedifferent=0
def updatetime():  #保持显示时间的更新
    global timestr  #timestr在后面日历中也要用来确定当天日期和闹钟
    timestr=time.localtime(time.time()-timedifferent)
    solardate=Solar.fromYmd(timestr.tm_year,timestr.tm_mon,timestr.tm_mday)
    lunardate=solardate.getLunar()
    printtime.setText(time.strftime('%H:%M:%S\n%Y年%m月%d日星期',timestr)+solardate.getWeekInChinese()+'\n农历'+str(lunardate)[5:]+'\n'+special_for_day(timestr.tm_year,timestr.tm_mon,timestr.tm_mday))
def updateweather():  #天气不需要一直更新,手动更新即可
    try:
        r=requests.get('http://www.weather.com.cn/weather/'+str(cities[choosecity])+'.shtml', timeout = 1)  
        r.raise_for_status()         
        r.encoding = r.apparent_encoding
        weatherdata=r.text
        index1=weatherdata.find(time.strftime('%m月%d日',timestr))
        if index1 == -1 :
            time_for_weather=time.localtime(time.mktime(timestr)-86400)
            index1=weatherdata.find(time.strftime('%m月%d日',time_for_weather))
            if index1 == -1 :
                printweather.setText('奇怪,日期对不上啊...\n试试看与网络时间同步?')
            else:
                index2=weatherdata.find('°C" />')
                printweather.setText(choosecity+'\n'+weatherdata[index1+14:index2+2])
        else:
            index2=weatherdata.find('°C" />')
            printweather.setText(choosecity+'\n'+weatherdata[index1+14:index2+2])
        #在查询天气数据时,我用了一个取巧的方法,在查阅中国天气网的天气网页源代码时,我发现当天的天气这一行开头必定是'X月X日X时',且其在整段代码中唯一出现,因此我只需定位当天日期这个关键词即可快速定位到天气这一行
        #但是有一个小问题,凌晨时网站上的当天天气数据还未更新,需要调取前一天的天气数据,不过这个属于正常现象.
    except:
        printweather.setText('查询天气数据失败QAQ\n请检查网络是否连接')
def manual_update_action():  #手动更新天气和日历的界面
    updateweather()
    calendar.updatecanlender(timestr)
def special_for_day(year,month,day):  #判断某天是否为节日或节气
    solardate=Solar.fromYmd(year,month,day)
    lunardate=solardate.getLunar()
    specials=[]
    if lunardate.getJieQi() != '':
        specials.append(lunardate.getJieQi())
    for festival in solardate.getFestivals():
        if festival in ['元旦节','劳动节','妇女节','圣诞节','儿童节','建党节','建军节','情人节','教师节'] or (festival == '国庆节' and year>=1949):
            specials.append(festival)
    for festival in lunardate.getFestivals():
        if festival in ['除夕','春节','元宵节','端午节','七夕节','中秋节','重阳节']:
            specials.append(festival)
    if year >=1952 and month == 10 and day == 25:
        specials.append('北航校庆日')
    if birthday != None and year >= birthday[0] and month == birthday[1] and  day == birthday[2]:
        specials.append('生日')
    return ' '.join(specials)
def update_notebook():  #将修改后的备忘录数据写回note.xlxs,同时将闹钟的计时器重新设定
    global alarmindex
    global alarmtitle
    global alarmnote  
    global alarm  #更新后的闹钟属性要能够传递出去
    wb = Workbook()
    ws = wb.active
    times_left=[]
    if notebook != ([]):
        for tip in notebook:
            ws.append(list(tip))
            time_left=time.mktime(time.strptime(f'{tip[0]} {tip[1]}', '%Y-%m-%d %H:%M'))-time.mktime(timestr)
            if time_left<-1:
                time_left = float('inf')
            times_left.append(time_left)
        if min(times_left) != float('inf'):
            alarmindex=times_left.index(min(times_left))
            alarmtitle=notebook[alarmindex][2]
            alarmnote=notebook[alarmindex][3]
            alarm=core.QTimer()
            alarm.setSingleShot(True)
            alarm.timeout.connect(alarm_time_out)
            alarm.start(int(min(times_left))*1000)
    while True:
        try:
            wb.save('note.xlsx')
            break
        except:
            widgets.QMessageBox.information(window, '写入备忘录文件时出错', 'CalenderZ出错了!\n这肯定不是程序的问题!\n绝对不是!\n(请关闭占用"note.xlxs"文件的应用)')
def alarm_time_out():  #闹钟,或者说定时器,响了
    newwindow=alarmwindow()
    newwindow.exec()
    notebook.pop(alarmindex)
    update_notebook()
def note_control_action():  #打开显示备忘录管理的窗口
    newindow = note_control_window()
    newindow.exec_()

class CalendarZ(widgets.QCalendarWidget):  #在pyqt5自带日历基础上改进其它功能
    def __init__(self, timestr,*args, **kwargs):
        super().__init__(*args, **kwargs)
        self.today = core.QDateTime.fromSecsSinceEpoch(int(time.mktime(timestr))).date()
    def paintCell(self, painter, rect, date):  #当前日期默认标红
        if date == self.today:
            painter.save()
            painter.fillRect(rect, core.Qt.red)
            painter.setPen(core.Qt.black)
            painter.drawText(rect, core.Qt.AlignCenter, str(date.day()))
            painter.restore()
        else:
            super().paintCell(painter, rect, date)
        year=date.year()
        month=date.month()
        day=date.day()
        special=special_for_day(year,month,day)
        painter.save()
        painter.setPen(core.Qt.blue)
        painter.drawText(rect, core.Qt.AlignBottom | core.Qt.AlignRight, special)
        painter.restore()
    def updatecanlender(self, timestr):  #改变日历的当天显示,也只需手动更新
        self.today = core.QDateTime.fromSecsSinceEpoch(int(time.mktime(timestr))).date()
        self.updateCells()
    def contextMenuEvent(self, event):  #对某一天的操作菜单
        menu = widgets.QMenu(self)
        menu.addAction('计算与今天天数差值',self.datedifference)
        menu.addAction('添加提醒',self.addnotes)
        menu.addAction('设为生日',self.addbirthday)
        menu.exec_(event.globalPos())
    def datedifference(self):  #计算与今天天数差值
        selected_date = self.selectedDate()
        difference = self.today.daysTo(selected_date)
        if difference > 0:
            message = f'所选日期在今天后{difference}天'
        elif difference <0:
            message = f'所选日期在今天前{-difference}天'
        else:
            message = f'所选日期为今天'
        widgets.QMessageBox.information(self, '结果', message)
    def addnotes(self):  #为日历中选定的某天添加提醒
        selected_date = self.selectedDate()
        newwindow = add_note_window(self)
        if newwindow.exec_() == widgets.QDialog.Accepted:
            title = newwindow.title_edit.text()
            note = newwindow.note_edit.toPlainText()
            hour = newwindow.hour_spinbox.value()
            minute = newwindow.minute_spinbox.value()
            time = f'{hour:02d}:{minute:02d}'
            notebook.append((selected_date.toString('yyyy-MM-dd'),time,title,note))
            update_notebook()
    def addbirthday(self):
        global birthday
        selected_date = self.selectedDate()
        birthday=(selected_date.year(),selected_date.month(),selected_date.day())
def city_choose_action(city):  #更换选择的城市
    global choosecity
    choosecity=city
    updateweather()
class add_note_window(widgets.QDialog):  #添加提醒的窗口
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle('添加提醒')
        self.resize(600,450)
        self.title_edit = widgets.QLineEdit()
        self.note_edit = widgets.QPlainTextEdit()
        self.note_edit.setMinimumHeight(100)
        self.note_edit.textChanged.connect(self.adjust_height)
        self.hour_spinbox = widgets.QSpinBox()
        self.hour_spinbox.setRange(0, 23)
        self.minute_spinbox = widgets.QSpinBox()
        self.minute_spinbox.setRange(0, 59)
        self.ok_button = widgets.QPushButton('确定')
        self.ok_button.clicked.connect(self.accept)
        self.cancel_button = widgets.QPushButton('取消')
        self.cancel_button.clicked.connect(self.reject)
        layout = widgets.QFormLayout(self)
        layout.addRow('标题：', self.title_edit)
        layout.addRow('内容：', self.note_edit)
        time_layout = widgets.QHBoxLayout()
        time_layout.addWidget(self.hour_spinbox)
        time_layout.addWidget(widgets.QLabel('时'))
        time_layout.addWidget(self.minute_spinbox)
        time_layout.addWidget(widgets.QLabel('分'))
        layout.addRow('具体时间：', time_layout)
        button_layout = widgets.QHBoxLayout()
        button_layout.addWidget(self.ok_button)
        button_layout.addWidget(self.cancel_button)
        layout.addRow(button_layout)
    def adjust_height(self):  # 根据内容自动调整输入框高度
        document_height = self.note_edit.document().size().height()
        margin = self.note_edit.contentsMargins().top() + self.note_edit.contentsMargins().bottom()
        new_height = int(document_height + margin)
        self.note_edit.setMinimumHeight(new_height)
class note_control_window(widgets.QDialog):  #显示备忘录管理的窗口
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle('备忘录管理')
        self.table = widgets.QTableWidget(0, 5)
        self.table.setHorizontalHeaderLabels(['日期', '时间', '标题', '内容', '操作'])
        for row in range(len(notebook)):
            self.table.insertRow(row)
            for column in range(4):
                self.table.setItem(row,column,widgets.QTableWidgetItem(notebook[row][column]))
            delete_button = widgets.QPushButton('删除')
            delete_button.clicked.connect(lambda _, row=row: self.delete(row))
            self.table.setCellWidget(row, 4, delete_button)
        self.save_button = widgets.QPushButton('保存')
        self.save_button.clicked.connect(self.save)
        self.cancel_button = widgets.QPushButton('取消')
        self.cancel_button.clicked.connect(self.reject)
        layout = widgets.QVBoxLayout(self)
        layout.addWidget(self.table)
        button_layout = widgets.QHBoxLayout()
        button_layout.addWidget(self.save_button)
        button_layout.addWidget(self.cancel_button)
        layout.addLayout(button_layout)
        self.resize(680, 450)
    def delete(self, row):  #删除按钮对应的操作,每次删完要重新连接以免对不上
        self.table.removeRow(row)
        for row in range(self.table.rowCount()):
            delete_button = self.table.cellWidget(row, 4)
            delete_button.clicked.disconnect()
            delete_button.clicked.connect(lambda _, row=row: self.delete(row))
    def save(self):
        notebook.clear()
        for row in range(self.table.rowCount()):
            date_str = self.table.item(row, 0).text()
            time_str = self.table.item(row, 1).text()
            title = self.table.item(row, 2).text()
            note = self.table.item(row, 3).text()
            notebook.append((date_str, time_str, title, note))
        update_notebook()
        self.accept()
class alarmwindow(widgets.QDialog):  #显示提醒的窗口
    def __init__(self, parent=None):
        super(alarmwindow, self).__init__(parent)
        self.setWindowTitle(alarmtitle)
        self.label = widgets.QLabel(alarmnote, self)
        self.label.setAlignment(core.Qt.AlignCenter)
        font = self.label.font()
        font.setPointSize(10)
        self.label.setFont(font)
        layout = widgets.QVBoxLayout(self)
        layout.addWidget(self.label)
        layout.addWidget(self.label)
        self.resize(400, 250)
        self.player=media.QMediaPlayer()
        url = core.QUrl.fromLocalFile('alarm.mp3')
        self.player.setMedia(media.QMediaContent(url))
        self.player.play()
    def closeEvent(self, event):
        self.player.stop()


#创建窗口
app = widgets.QApplication(sys.argv)
window = widgets.QMainWindow()
window.resize(800, 600)
window.setWindowTitle('CanlenderZ')

#左上角日期时间等的显示设置
printtime = widgets.QLabel(parent=window)
font=printtime.font()
font.setPointSize(15)
printtime.setFont(font)
updatetime()

#右上角天气查询功能的显示设置
printweather = widgets.QLabel(parent=window)
font=printweather.font()
font.setPointSize(15)
printweather.setFont(font)
updateweather()

#下方为日历的显示设置
calendar = CalendarZ(timestr)
calendar.setMinimumDate(core.QDate(1900, 1, 1))
calendar.setMaximumDate(core.QDate(2100, 12, 31))

#备忘录的导入
try:
    wb = load('note.xlsx')
    ws = wb.active
    notebook=list(ws.iter_rows(min_row=1, values_only=True))
    times_left=[]
    if notebook != [(None,)]:
        for tipindex in range(len(notebook)):  #读取备忘录内容时立即为当前最早的提醒设定一个闹钟,在未修改备忘录内容时,这一闹钟的倒计时不变
            if len(notebook[tipindex]) == 3:  #某些人写备忘录不带内容导致读取出错,必须修正一下
                notebook[tipindex]+=(None,)
            time_left=time.mktime(time.strptime(f'{notebook[tipindex][0]} {notebook[tipindex][1]}', '%Y-%m-%d %H:%M'))-time.mktime(timestr)
            if time_left<-1:
                time_left = float('inf')
            times_left.append(time_left)
        if min(times_left) != float('inf'):
            alarmindex=times_left.index(min(times_left))
            alarmtitle=notebook[alarmindex][2]
            alarmnote=notebook[alarmindex][3]
            alarm=core.QTimer()
            alarm.setSingleShot(True)
            alarm.timeout.connect(alarm_time_out)
            alarm.start(int(min(times_left)*1000))
    else:
        notebook=[]
except:
    widgets.QMessageBox.information(window, '读取备忘录文件时出错', 'CalenderZ出错了!\n这肯定不是程序的问题!\n绝对不是!\n(请检查"note.xlxs"文件是否存在)')
    sys.exit(app.exec_())

#页面排版
hlayout = widgets.QHBoxLayout()
hlayout.addWidget(printtime)
hlayout.addStretch()
hlayout.addWidget(printweather)
vlayout = widgets.QVBoxLayout()
vlayout.addLayout(hlayout)
vlayout.addWidget(calendar)
centralWidget = widgets.QWidget()
centralWidget.setLayout(vlayout)
window.setCentralWidget(centralWidget)

#保持页面的定时刷新,按114ms为周期更新时钟(只是为了减小更新周期读秒的误差,数字本身没有其它用处)
clock = core.QTimer()
clock.timeout.connect(updatetime)
clock.start(114)

#创建菜单
menubar = window.menuBar()
mode=menubar.addMenu('功能')
manualupdate=widgets.QAction('手动更新天气和日历',window)
manualupdate.triggered.connect(manual_update_action)
mode.addAction(manualupdate)
returntoday=widgets.QAction('回到今天',window)
returntoday.triggered.connect(lambda: calendar.setSelectedDate(core.QDate.currentDate()))
mode.addAction(returntoday)
notecontrol=widgets.QAction('备忘录管理',window)
notecontrol.triggered.connect(note_control_action)
mode.addAction(notecontrol)
action=menubar.addMenu('设置')
timecheck=action.addMenu('时间同步')
systemtime=widgets.QAction("与系统时间同步",window)
systemtime.triggered.connect(systemtimechange)
timecheck.addAction(systemtime)
nettime=widgets.QAction("与网络时间同步",window)
nettime.triggered.connect(nettimechange)
timecheck.addAction(nettime)
citieschoose=action.addMenu('选择城市')
for city in cities.keys():
    citychoose=widgets.QAction(city, window)
    citychoose.triggered.connect(lambda checked, c=city: city_choose_action(c))
    citieschoose.addAction(citychoose)
#界面显示与关闭
window.show()
sys.exit(app.exec_())
